import os
import requests
import pandas as pd
from tqdm import tqdm
from copy import deepcopy
from utils import get_district_name
from utils import get_application_key
    
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class WebQueryClient(object):
    def __init__(self, 
                query_params:dict={}, 
                decode_key:dict={}, 
                verbose:int=0, 
                **kwargs
                ):
        
        self.decode_key = decode_key
        self.query_params = {}
        for k, v in query_params.items():
            if k == "cortarName":
                continue
            if len(v) > 1:
                value = ""
                for sub_v in v:
                    value+=str(sub_v)
                    value+=":"
                self.query_params[str(k)] = str(value[:-1])
                
            else:
                self.query_params[str(k)] = str(v[0])
                
        self.basic_param_key = [ 
            'realEstateType',
            'tradeType',
            'cortarName',
            'PriceMin',
            'PriceMax',
            'areaMin',
            'areaMax',
            ]
                
        self.verbose = verbose
                
        self.headers = {
            'accept': '*/*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpZCI6IlJFQUxFU1RBVEUiLCJpYXQiOjE3NDg2OTk5ODgsImV4cCI6MTc0ODcxMDc4OH0.M4fq7xuiqtR7HPGLhSgMABDVyXveDYJvezKnlBmQfwc',
            'priority': 'u=1, i',
            'referer': 'https://new.land.naver.com/offices?ms=37.3843058,127.2867088,15&a=TJ&e=RETAIL',
            'sec-ch-ua': '"Chromium";v="136", "Google Chrome";v="136", "Not.A/Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36',
            # 'cookie': 'nhn.realestate.article.ipaddress_city=1100000000; NAC=7d6sBoAZJBfrB; NACT=1; _fwb=85YUStmxwtN5ssGduSQIUR.1748699444357; landHomeFlashUseYn=Y; NNB=UUMDIURQBE5WQ; SRT30=1748699440; nhn.realestate.article.trade_type_cd=A1; realestate.beta.lastclick.cortar=4100000000; _fwb=85YUStmxwtN5ssGduSQIUR.1748699444357; _ga=GA1.1.1027182781.1748699461; _ga_451MFZ9CFM=GS2.1.s1748699461$o1$g0$t1748699462$j59$l0$h0; SHOW_FIN_BADGE=Y; BNB_FINANCE_HOME_TOOLTIP_ESTATE=true; nhn.realestate.article.rlet_type_cd=Z04; REALESTATE=Sat%20May%2031%202025%2022%3A59%3A48%20GMT%2B0900%20(Korean%20Standard%20Time); SRT5=1748700431; BUC=GyJMIt_y4lzbPeOPi0sIOfq1G5Y3mvf55uIua0AnsJg=',
        }
        
        # preprocessing function
        self.kwargs = kwargs
        self.add_get_address = self.kwargs.get('address', False)
        self.add_get_ETA = self.kwargs.get('ETA', False)
        self.add_get_map_link = self.kwargs.get('map_link', False)
        self.add_get_rlet_link = self.kwargs.get('rlet_link', False)
        
        if (self.add_get_address or 
            self.add_get_ETA or
            self.add_get_map_link
            ):
            application_key = get_application_key()
            self.headers_post = {
                'X-NCP-APIGW-API-KEY-ID': application_key['client_id'],
                'X-NCP-APIGW-API-KEY': application_key['client_secret'],
            }
            
    def print(self, level:int, info:str):
        if self.verbose >= level:
            print(info)
        
    def get_basic_data(self, district_code):
        self.print(2, f"[I] Start with getting basic data.")
        
        basic_url = "https://new.land.naver.com/api/articles"
        total_data = []
        
        self.print(2, f"[I] Define total_data: {total_data}.")
        
        params = {}
        for key, value in self.query_params.items():
            if key in self.basic_param_key:
                params[key] = value
                
        self.print(2, f"[I] Copy query_params: {params}.")
        
        params['cortarNo'] = district_code
        params['page'] = 1
        self.print(2, f"[I] Add cortarNo & page at params: {params}.")
        continue_ = True
        
        while continue_:
            self.print(2, f"[I] Start with request information from {basic_url}.")
            response = requests.get(
                url=basic_url, 
                params=params, 
                headers=self.headers,
                )
            self.print(2, f"[I] Request URL: {response.url}.")
            self.print(2, f"[I] Response: {response}.")
            
            # check response status
            if response.status_code == 200:
                data_dic = response.json()
                self.print(2, f"[I] Number of Received data: {len(data_dic['articleList'])}.")
                
                total_data += data_dic['articleList']
                
                # check there is more data
                if not data_dic['isMoreData']:
                    self.print(2, f"[I] There is no more basic data.")
                    break
                
                params['page'] += 1
                
            else:
                print(f"Failed to retrieve data: {response.status_code}")
                break
                
        return total_data
    
    def get_detail_data(self, basic_data):
        self.print(2, f"[I] Start with getting detail data.")
        
        detail_url = "https://new.land.naver.com/api/articles"
        total_detail_data = []
        
        self.print(2, f"[I] Define total_detail_data: {total_detail_data}.")
        
        for i, data in enumerate(basic_data):
            self.print(2, f"[I] Define url by article number. [{i+1}/{len(basic_data)}]")
            url=f"{detail_url}/{data['articleNo']}"
            
            self.print(2, f"[I] Start with request information from {url}.")
            response = requests.get(
                url=url,
                headers=self.headers,
                )
            self.print(2, f"[I] Request URL: {response.url}.")
            
            # check response status
            if response.status_code == 200:
                self.print(2, f"[I] Response: {response}.")
                data_dic = response.json()
                
                total_detail_data.append(data_dic)
                
        return total_detail_data
            
    def find_all_values_with_keys(self, dict_, current_path=None):
        if current_path is None:
            current_path = []
        values_with_keys = []
        for key, value in dict_.items():
            new_path = current_path + [key]
            if isinstance(value, dict):
                values_with_keys.extend(self.find_all_values_with_keys(value, new_path))
            else:
                values_with_keys.append((new_path, value))
        return values_with_keys
            
    def decode_to_dataframe(self, basic_data, detail_data):
        self.print(2, f"[I] Start with decoding to dataframe.")
        basic_columns = self.decode_key["basic"].values()
        
        detail_keys, detail_columns = zip(*self.find_all_values_with_keys(self.decode_key["detail"]))
        columns = list(basic_columns) + list(detail_columns)
        
        decode_dict = {key:[] for key in columns}
        
        for i, basic_ in enumerate(basic_data):
            self.print(2, f"[I] Decode data [{i+1}/{len(basic_data)}]")
            values = [basic_[k] if k in basic_.keys() else None for k in self.decode_key["basic"].keys()]
            cols = deepcopy(columns)
            
            for keys in detail_keys:
                value = deepcopy(detail_data[i])
                
                for key in keys[:-1]:
                    value = deepcopy(value[key])
                    
                last_key = keys[-1]
                
                if last_key in ["latitude", "longitude"]:
                    try:
                        value = value[last_key]
                    
                        basic_value = values[list(self.decode_key['basic'].keys()).index(last_key)]
                    
                        if (basic_value == '0') & (not value[last_key] == '0'):
                            values[list(self.decode_key['basic'].keys()).index(last_key)] = value[last_key]
                            self.print(2, f"[I] {last_key} update : {basic_value}->{value[last_key]}.")
                        
                    except:
                        pass
                    
                    value = 'pass'
                    
                else:
                    try:
                        value = value[last_key]
                    except:
                        value = '-'
                        self.print(2, f"[I] There is no {last_key} information.")
                        
                if value == 'pass':
                    cols.pop(len(values))
                else:
                    values += [value]
                    
            for i, col in enumerate(cols):
                decode_dict[col].append(values[i])
        
        # delete unused key
        del_keys = []
        for k, v in decode_dict.items():
            if len(v) == 0:
                del_keys.append(k)
        for k in del_keys:
            del decode_dict[k]
                
        df = pd.DataFrame(decode_dict)
        
        return df
    
    def request_address(self, lat, lng):
        self.print(2, f"[I] Start with getting address. [lat:{lat}, lon:{lng}]")
        add = ''
        orders = "addr"     # 지번
        
        reverse_gc_url = "https://naveropenapi.apigw.ntruss.com/map-reversegeocode/v2/gc"
        url = f"{reverse_gc_url}?coords={lng},{lat}&output=json&orders={orders}"
        
        self.print(2, f"[I] Start with request information from {url}.")
        response = requests.get(
                    url,
                    headers=self.headers_post,
                    )
                
        
        # 응답 상태 코드 확인
        if response.status_code == 200:
            self.print(2, f"[I] Response: {response}.")
            # JSON 데이터 파싱
            data_dic = response.json()
            if len(data_dic['results']) != 0:
            
                add += data_dic['results'][0]['region']['area3']['name'] + ' '
                area4 = data_dic['results'][0]['region']['area4']['name']
                if area4 != '':
                    add += area4 + ' '
                
                type_ = data_dic['results'][0]['land']['type']
                if type_ == '2':
                    add += '산 '
                
                add += data_dic['results'][0]['land']['number1']
                
                num2 = data_dic['results'][0]['land']['number2']
                if num2 != '':
                    add += '-' + num2
                
                self.print(2, f"[I] Get Address: {add}.")
                    
            else:
                print(f"Failed to retrieve data: {response.status_code}, {response.text}")
                
        return add
    
        
    def get_address(self, df):
        df['주소'] = ''
        
        # orders = "roadaddr" # 도로명
        print(f"[I] 토지 주소 변환 중...")
        for i in tqdm(range(len(df))):
            self.print(2, f"[I] Getting address [{i+1}/{len(df)}]")
            
            row = df.iloc[i]
            lat = float(row['위도'])
            lng = float(row['경도'])
            
            if (lat==0) or (lng==0):
                self.print(2, f"[I] There is no latitude & longitudinal data: ArticleNo={row['매물번호']}.")
                add = '-'
            else:
                add = self.request_address(lat, lng)
            df.iloc[i, df.columns.get_loc("주소")] = add
        
        return df
    
    def request_ETA(self, start_lat, start_lng, goal_lat, goal_lng):
        eta_m = ''
        params = {
            "start": f"{start_lng},{start_lat}",
            "goal": f"{goal_lng},{goal_lat}",
            "option": "trafast",     # 빠른 경로를 위한 옵션
            # "option": "traoptimal",  # 최적 경로를 위한 옵션
        }
        driving_url = "https://naveropenapi.apigw.ntruss.com/map-direction/v1/driving"
        
        response = requests.get(
                    driving_url, 
                    params=params,
                    headers=self.headers_post,
                    )
        
        # 응답 상태 코드 확인
        if response.status_code == 200:
            # JSON 데이터 파싱
            data_dic = response.json()
            
            if data_dic['code'] == 0:
                request_time = data_dic['currentDateTime']
                eta_ms = data_dic['route'][params['option']][0]['summary']['duration']
                eta_m = (eta_ms // 1000) // 60
            else:
                raise ConnectionError(f"[E] Failed to retrieve ETA data: {response.status_code}, {response.text}")
                
        else:
            raise ConnectionError(f"[E] Failed to retrieve ETA data: {response.status_code}, {response.text}")
                
        return eta_m, request_time
        
    def get_ETA(self, df, start="강남역"):
        if start == "강남역":
            start_lat = 37.498716
            start_lng = 127.027021
        else:
            raise NotImplementedError("[E] Starting points other than Gangnam Station are not implemented or provided.")
        
        df['소요시간(분)'] = ''
        df['소요시간 탐색 시각'] = ''
        
        print(f"[I] 소요시간 (from {start}) 계산 중...")
        for i in tqdm(range(len(df))):
            
            row = df.iloc[i]
            lat = float(row['위도'])
            lng = float(row['경도'])
            
            if (lat==0) or (lng==0):
                self.print(2, f"[I] There is no latitude & longitudinal data: ArticleNo={row['매물번호']}.")
                eta_m, request_time = '-', '-'
            else:
                eta_m, request_time = self.request_ETA(start_lat, start_lng, lat, lng)
            
            df.iloc[i, df.columns.get_loc("소요시간(분)")] = eta_m
            df.iloc[i, df.columns.get_loc("소요시간 탐색 시각")] = request_time
        
        return df
    
    def get_map_link(self, df):
        map_url = "https://map.naver.com/p/search/"
        df["네이버지도"] = ''
        
        print(f"[I] 네이버지도 링크 생성 중...")
        for i in tqdm(range(len(df))):
            
            row = df.iloc[i]
            add = str(row['주소']).split(" ")
            url = map_url
            for add in str(row['주소']).split(" "):
                url += add + "%20"
            
            df.iloc[i, df.columns.get_loc("네이버지도")] = f'=HYPERLINK("{url[:-3]}", "링크")'
        
        return df

    def get_rlet_link(self, df):
        rlet_url = "https://fin.land.naver.com/articles"
        df["네이버부동산"] = ''
        
        print(f"[I] 네이버부동산 링크 생성 중...")
        for i in tqdm(range(len(df))):
            
            row = df.iloc[i]
            
            url = f"{rlet_url}/{row['매물번호']}"
            
            df.iloc[i, df.columns.get_loc("네이버부동산")] = f'=HYPERLINK("{url}", "링크")'
        
        return df
        
    def post_processing(self, df):
        if self.add_get_ETA:
            df = self.get_ETA(df)
        if self.add_get_address:
            df = self.get_address(df)
        if self.add_get_map_link:
            if not self.add_get_address:
                raise KeyError("[E] 링크를 추가하려면 주소 기능을 활성화하세요.")
            df = self.get_map_link(df)
        if self.add_get_rlet_link:
            df = self.get_rlet_link(df)
        
        return df
    
    def modify_columns_order(self, df):
        if '설명' in df.columns:
            old_columns = list(df.columns)
            old_columns.pop(old_columns.index('설명'))
            new_columns = old_columns + ['설명']
            df = df[new_columns]
        
        return df
    
    def fit_width(self, excel_path):
        # 열 너비 자동 조정
        wb = load_workbook(excel_path)
        ws = wb.active
        
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)
            
            # 열 이름 길이 기준 초기값
            header = ws.cell(row=1, column=col_idx).value
            max_length = len(str(header))*1.5 if header else 0

            if header in ['네이버지도', '네이버부동산']:
                continue
            
            # 데이터 행 길이 비교
            for cell in column_cells[1:]:  # 첫 번째는 헤더이므로 생략 가능
                try:
                    cell_value = str(cell.value)
                    if cell_value is not None:
                        max_length = max(max_length, len(cell_value)*1.5)
                except:
                    pass
            
            # 여유 공간 포함해서 설정
            ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(excel_path)
                
        
    def get_n_save_data(self, district_codes, save_dir=os.path.join(os.path.expanduser("~"), "Downloads", "LandScout")):
        for i, code in enumerate(district_codes):
            
            code_name = get_district_name(code)
            self.print(0, f"[I] ({i+1}/{len(district_codes)}) {code_name} 작업 중...")
            self.print(2, f"[I] {code_name} code is {code}.")
            
            # get basic data
            basic_data = self.get_basic_data(code)
            if len(basic_data) < 1:
                print(f"[I] {code_name}: There is no valid data.")
                continue
            
            # get detail data
            detail_data = self.get_detail_data(basic_data)
            if len(basic_data) != len(detail_data):
                print(f"[I] [{code_name}] Failed to retrieve detailed information. # of basic info: {len(basic_data)} / # of detail info: {len(detail_data)} ")
                continue
                
            df = self.decode_to_dataframe(basic_data, detail_data)
            print(f"[I] {code_name} 토지 검색 완료 (총 {len(df)} 도출).")
            
            # post processing
            if sum(self.kwargs.values()) > 0:
                df = self.post_processing(df)
            
            if not os.path.isdir(save_dir):
                os.makedirs(save_dir)
                
            df = self.modify_columns_order(df)
            
            save_path = os.path.join(save_dir, f'{code_name}.xlsx')
            df.to_excel(save_path, index=False)
            
            self.fit_width(save_path)
            
            print(f"[I] {code_name} 저장 완료.")